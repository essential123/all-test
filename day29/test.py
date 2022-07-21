import re

import pandas
from openpyxl import Workbook, load_workbook

#
# wb = Workbook()
# wb1 = wb.create_sheet('天上人间', 0)
# wb2 = wb.create_sheet('红色浪漫')
# wb3 = wb.create_sheet('无人问津')
#
# wb1.append(['name', 'age', 'hobby'])
# wb1.append(['T', 18, 'game'])
# wb1.append(['A', 18, 'read'])
# wb1.append(['0', 18, 'music'])
#
# wb.save(r'bar.xlsx')

# from openpyxl import load_workbook
#
# wb = load_workbook(r'bar.xlsx')
# print(wb.sheetnames) # 查看excel文件中所有工作簿名称 ['天上人间', 'Sheet', '红色浪漫', '无人问津']
# wb1 = wb['天上人间']
# print(wb1) # <Worksheet "天上人间">
# print(wb1.max_row) # 显示工作簿最多有几行内容 4
# print(wb1.max_column) # 显示工作簿最多有几列内容 3
# print(wb1['A1']) # <Cell '天上人间'.A1>
# print(wb1['A1'].value) # 第一种取值方式 name
# print(wb1.cell(row=1,column=1).value) # 第二种取值方式 name
# print(wb1.rows) # 生成器节省内存 <generator object Worksheet._cells_by_row at 0x000000000B3E5F90>
# for i in wb1.rows:
#     print(i)
# # (<Cell '天上人间'.A1>, <Cell '天上人间'.B1>, <Cell '天上人间'.C1>)
# # (<Cell '天上人间'.A2>, <Cell '天上人间'.B2>, <Cell '天上人间'.C2>)
# # (<Cell '天上人间'.A3>, <Cell '天上人间'.B3>, <Cell '天上人间'.C3>)
# # (<Cell '天上人间'.A4>, <Cell '天上人间'.B4>, <Cell '天上人间'.C4>)
# for i in wb1.rows:
#     print([[j.value for j in i]])
# # [['name', 'age', 'hobby']]
# # [['T', 18, 'game']]
# # [['A', 18, 'read']]
# # [['0', 18, 'music']]

# import pandas
#
# d = {
#     '公司名称': ['老男孩','老伙计','老北鼻'],
#     '公司地址': ['上海', '深圳', '杭州'],
#     '公司电话': [120, 130, 996],
# }
#
# df = pandas.DataFrame(d)
# df.to_excel(r'company.xlsx')

# import requests
# res = requests.get('https://sh.lianjia.com/ershoufang/')
# with open('链家.html','wb')as f:
#     f.write(res.content)

#
# with open('链家.html','r',encoding='utf8')as f :
#     data = f.read()
#
# house_title_list = re.findall('<a class="" .*?" target="_blank" data-log_index=".*?"  data-el="ershoufang".*?" data-is_focus="" data-sl="">(.*?)</a>',data)
# # print(house_title_list)
#
# house_name_list = re.findall('<a href=".*?" target="_blank" data-log_index=".*?" data-el="region">(.*?) </a>',data)
# # print(house_name_list)
#
# house_addr_list = re.findall('</a>   -  <a href=".*?" target="_blank">(.*?)</a>',data)
# # print(house_addr_list)
#
# house_info_list = re.findall('<div class="houseInfo"><span class="houseIcon"></span>(.*?)</div>',data)
# # print(house_info_list)
#
# house_collection_list = re.findall('<div class="followInfo"><span class="starIcon"></span>(.*?)</div>',data)
# # print(house_collection_list)
#
# house_total_price = re.findall('<div class="totalPrice totalPrice2"><i> </i><span class="">(.*?)</span><i>万</i></div>',data)
# # print(house_total_price)
#
# house_unit_price = re.findall('<div class="unitPrice" data-hid=".*?" data-rid=".*?" data-price=".*?"><span>(.*?)/平</span></div>',data)
# # print(house_unit_price)
#
# house_dic = {
#     '房屋标题': house_title_list,
#     '小区名称': house_name_list,
#     '所在街道': house_addr_list,
#     '具体信息': house_info_list,
#     '其他信息': house_collection_list,
#     '房屋总价': house_total_price,
#     '房屋单价': house_unit_price
# }
#
# df = pandas.DataFrame(house_dic)
# df.to_excel(r'链家.xlsx')

# import random
#
# code = ''
#
# for i in range(5):
#     random_int = str(random.randint(0, 9))
#     random_lower = chr(random.randint(97, 120))
#     random_upper = chr(random.randint(65, 90))
#     temp = random.choice([random_int, random_lower, random_upper])
#     code += temp
# print(code)

import logging

file_handler = logging.FileHandler(filename='x1.log', mode='a', encoding='utf-8', )
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s -%(module)s:  %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S %p',
    handlers=[file_handler, ],
    level=logging.ERROR
)

logging.error('我不好!!!')






















